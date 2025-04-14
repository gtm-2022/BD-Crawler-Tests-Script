import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import datetime
from concurrent.futures import ThreadPoolExecutor

def get_search_data(root, term, max_pages,timeout):
    try:
        extracted_data = []
        print(f"Processing term: {term}")
        for page in range(1, max_pages + 1):  # Loop through multiple pages
            print(f"{term} - Page {page}")
            response = requests.get(f"http://localhost:5000/api/search?rootdomain={root}&term={term}&page={page}",timeout=timeout)
            
            count = 1
            while response.status_code != 200 and count < 3:  # Retry up to 3 times
                response = requests.get(f"http://localhost:5000/api/search?rootdomain={root}&term={term}&page={page}",timeout=timeout*(count+1))
                count += 1

            if response.status_code == 200:
                data = response.json()
                print(f"{term} - Page {page} succeeded after {count} attempts")
                
                search_items = data.get("searchItems", []) or []
                for index, item in enumerate(search_items, start=1):
                    # Safely get seller_sku and sku_entry with default empty dicts
                    seller_sku = item.get("sellerSku", {}) or {}
                    sku_entry = seller_sku.get("skuEntry", {}) or {}
                    
                    # Create a dictionary with safe defaults for all fields
                    extracted_data.append({
                        "statuscode": response.status_code,
                        "error_message": None,
                        "search_term": term,
                        "page": page,
                        "rank": (page - 1) * len(search_items) + index,
                        "title": item.get("title", ""),
                        "brand": item.get("brand", ""),
                        "price": item.get("price", ""),
                        "url": item.get("url", ""),
                        "sku": item.get("sku", ""),
                        "rootdomain": item.get("rootdomain", ""),
                        "average_customer_review": item.get("averageCustomerReview", ""),
                        "number_of_customer_reviews": item.get("numberOfCustomerReviews", ""),
                        "number_of_customer_ratings": sku_entry.get("numberOfCustomerRatings", ""),
                        "mpn": item.get("mpn", ""),
                        "is_sponsored": item.get("isSponsored", ""),
                        "promo_text": item.get("promoText", ""),
                        "shipping_type": item.get("shippingType", ""),
                        "get_it_by": item.get("getItBy", ""),
                        "number_of_favorites": item.get("numberOfFavorites", ""),
                        "list_price": item.get("listPrice", ""),
                        "open_box_price": item.get("openBoxPrice", ""),
                        "bestseller_text": item.get("bestsellerText", ""),
                        "quantity_sold": item.get("quantitySold", ""),
                        "description": sku_entry.get("description", ""),
                        "image_url": sku_entry.get("imageUrl", ""),
                        "upc": sku_entry.get("upc", ""),
                        "seller_id": seller_sku.get("sellerId", ""),
                        "timestamp": datetime.datetime.now().isoformat()
                    })
            else:
                error_message = response.text  # Capture error message from response

                extracted_data.append({
                    "statuscode": response.status_code,
                    "error_message": error_message,
                    "search_term": term,
                    "page": page,
                    "rank": None,
                    "title": None,
                    "brand": None,
                    "price": None,
                    "url": None,
                    "sku": None,
                    "rootdomain": None,
                    "average_customer_review": None,
                    "number_of_customer_reviews": None,
                    "number_of_customer_ratings": None,  # New field added
                    "mpn": None,
                    "is_sponsored": None,
                    "promo_text": None,
                    "shipping_type": None,
                    "get_it_by": None,
                    "number_of_favorites": None,
                    "list_price": None,
                    "open_box_price": None,
                    "bestseller_text": None,
                    "quantity_sold": None,
                    "description": None,
                    "image_url": None,
                    "upc": None,
                    "seller_id": None,
                    "timestamp": datetime.datetime.now().isoformat()
                })

                print(f"Failed to fetch data for term '{term}' on page {page}, Status Code: {response.status_code}")
        
        return extracted_data
    except requests.RequestException as e:
        print(f"Error fetching data for term '{term}': {e}")
        return []
    except Exception as e:
        print(f"Unexpected error processing term '{term}': {e}")
        return []

def run_searches_in_threads(workers,root, terms, max_pages,timeout):
    extracted_data = []
    with ThreadPoolExecutor(max_workers=workers) as executor:  # Using 3 worker threads
        futures = {executor.submit(get_search_data, root, term, max_pages,timeout): term for term in terms}
        
        for future in futures:
            term = futures[future]
            try:
                result = future.result()
                extracted_data.extend(result)
            except Exception as e:
                print(f"Error processing term '{term}': {e}")
    
    return extracted_data

def write_data_to_excel(output_file, data_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Search Data"
    
    headers = [
        "statuscode"," error_message","search_term", "page", "rank", "title", "brand", "price", "url", "sku","duplicate_sku", "rootdomain", "average_customer_review",
        "number_of_customer_reviews", "number_of_customer_ratings", "mpn", "is_sponsored", "promo_text", "shipping_type", "get_it_by",
        "number_of_favorites", "list_price", "open_box_price", "bestseller_text", "quantity_sold", 
        "description", "image_url", "upc", "seller_id", "timestamp"
    ]
    ws.append(headers)

    sku_column_idx = headers.index("sku") + 1  # SKU column index (1-based)
    duplicate_sku_idx = headers.index("duplicate_sku") + 1  # New duplicate column index
    
    sku_counts = {}

    
    # for data in data_list:
    #     ws.append([data.get(field, "") for field in headers])
    
    for data in data_list:
        row = []
        for field in headers:
            value = data.get(field, "")

            # Convert lists or dictionaries to strings
            if isinstance(value, (list, dict)):
                value = str(value)

            row.append(value)
        ws.append(row)

    # Save before applying styles to avoid memory issues
    wb.save(output_file)

     # Track SKUs for duplicate marking
    sku_value = str(data.get("sku", "")).strip()
    if sku_value:
        sku_counts[sku_value] = sku_counts.get(sku_value, []) + [ws.max_row]

    # ✅ Second pass: Mark duplicate SKUs
    for sku, rows in sku_counts.items():
        is_duplicate = len(rows) > 1  # If SKU appears more than once
        for row_idx in rows:
            ws.cell(row=row_idx, column=duplicate_sku_idx, value="True" if is_duplicate else "False")

    # ✅ Apply highlighting (Red for duplicates)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=duplicate_sku_idx)
        if cell.value == "True":  # Highlight only True values
            cell.fill = red_fill


    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Highlight columns where all values are the same (Red)
    for col_idx, header in enumerate(headers, start=1):
        column_values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(2, ws.max_row + 1)]

        # Remove None and empty strings
        non_empty_values = [val for val in column_values if val not in [None, ""]]

        if len(set(non_empty_values)) == 1 and len(non_empty_values) > 0:
            ws.cell(row=1, column=col_idx).fill = red_fill  # Highlight header in red
        else:
            continue

    # Highlight blank cells (Yellow)
    for col_idx, header in enumerate(headers, start=1):
        is_column_blank = True
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value in [None, "" , False,"False",0,"0","Unknown","unknown"]:  # Check if the cell is blank
                cell.fill = highlight_fill  # Apply yellow fill
            else:
                is_column_blank = False

        for row_idx in range(2, ws.max_row + 1):
            for img_col in ["url","image_url"]:
                col_idx = headers.index(img_col) + 1  # Find the column index
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and not str(cell.value).startswith("http"):  # Check if URL is invalid
                    cell.fill = highlight_fill
          
            for img_col in ["statuscode"]:
                col_idx = headers.index(img_col) + 1  # Find the column index
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and not str(cell.value).startswith("200"):  # Check if URL is invalid
                    cell.fill = red_fill  
    # # Highlight duplicate SKUs in red
    # for sku, rows in sku_counts.items():
    #     if len(rows) > 1:  # If SKU appears more than once
    #         for row_idx in rows:
    #             ws.cell(row=row_idx, column=sku_column_idx).fill = red_fill

    #     # If entire column is blank, highlight the header as well
    #     if is_column_blank:
    #         ws.cell(row=1, column=col_idx).fill = highlight_fill

    # Save the workbook after applying styles
    wb.save(output_file)
    print(f"Data saved to {output_file}")

def main(workers,root, terms, max_pages,timeout):
    output_file = fr"Demo_Search_{root.replace('.','_').replace('/','_')}.xlsx"
    search_data = run_searches_in_threads(workers,root, terms, max_pages,timeout)
    if search_data:
        write_data_to_excel(output_file, search_data)
    else:
        print("No data extracted.")

if __name__ == "__main__":
    timeout=5000
    workers=10
    root =  "hp.com/au"
    terms = ["laptop", "printers", "headphone", "camera", "usb", "monitor", "vr", "keyboard", "mouse", "charger"]   # Add multiple terms here 
    max_pages = 4  # Number of pages per term
    main(workers,root, terms, max_pages,timeout)
