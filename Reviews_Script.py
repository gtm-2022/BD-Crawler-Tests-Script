import requests
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

def fetch_reviews(root_domain, sku, max_pages, retry_attempts, timeout):
    all_reviews = []

    for page in range(1, max_pages + 1):
        print(f"Fetching {sku} - Page {page}...")
        url = f"http://localhost:5000/api/review?rootdomain={root_domain}&sku={sku}&page={page}"

        for attempt in range(1, retry_attempts + 1):  # Retry logic
            try:
                response = requests.get(url, timeout=timeout)
                status_code = response.status_code

                if status_code == 200:
                    data = response.json()
                    review_items = data.get("reviewItems", [])

                    if not review_items:
                        print(f"No more reviews for {sku} on page {page}. Stopping.")
                        break

                    for review in review_items:
                        all_reviews.append({
                            "statuscode": status_code,
                            "error_message": None,
                            "sku": sku,
                            "page": page,
                            "variantSku": review.get("variantSku"), 
                            "reviewId": review.get("reviewId"), 
                            "author": review.get("author"), 
                            "rating": review.get("rating"), 
                            "date": review.get("date"), 
                            "purchasedDate": review.get("purchasedDate"), 
                            "location": review.get("location"), 
                            "attributes": json.dumps(review.get("attributes", "")),  
                            "title": review.get("title"),
                            "text": review.get("text"),
                            "productName": review.get("productName"),  
                            "recommendedReview": review.get("recommendedReview"), 
                            "productHasBeenTried": review.get("productHasBeenTried"), 
                            "brandResponse": review.get("brandResponse"), 
                            "syndicated": review.get("syndicated"),
                            "program": review.get("program"), 
                            "link": review.get("link"), 
                            "reviewImagesUrl": json.dumps(review.get("reviewImagesUrl", "")),  
                            "sellerId": review.get("sellerId"),
                            "timestamp": datetime.datetime.now().isoformat()
                        })
                    break  # Exit retry loop on success
                else:
                    print(f"Attempt {attempt}: Status code {status_code} for {sku} - Page {page}")
                    if attempt == retry_attempts:
                        all_reviews.append({
                            "statuscode": status_code,
                            "error_message": response.text,
                            "sku": sku,
                            "page": page
                        })
            except requests.exceptions.RequestException as e:
                print(f"Attempt {attempt}: Error fetching {sku} - Page {page}: {e}")
                if attempt == retry_attempts:
                    all_reviews.append({
                        "statuscode": "Request Failed",
                        "error_message": str(e),
                        "sku": sku,
                        "page": page
                    })

    return all_reviews

def save_to_excel(reviews, output_path):
    if not reviews:
        print("No reviews to save. Skipping Excel file creation.")
        return

    fields = ["statuscode", "error_message", "sku", "page", "variantSku", "reviewId", "author", "rating", "date", "purchasedDate", 
              "location", "attributes", "title", "text", "productName", "recommendedReview", "productHasBeenTried", "brandResponse", 
              "syndicated", "program", "link", "reviewImagesUrl", "sellerId", "timestamp"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews Data"
    ws.append(fields)

    review_id_idx = fields.index("reviewId") + 1  # Review ID column index (1-based)
    review_counts = {}

    for review in reviews:
        row_data = []
        for field in fields:
            value = review.get(field, "")
            if isinstance(value, (list, dict)):
                value = json.dumps(value)  
            row_data.append(value)
        ws.append(row_data)

    for review in reviews:
        review_id = str(review.get("reviewId", "")).strip()
        if review_id:
            review_counts[review_id] = review_counts.get(review_id, []) + [ws.max_row]


    # ✅ Apply Formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for duplicates
  
    # ✅ Highlight duplicate Review IDs
    for review_id, rows in review_counts.items():
        if len(rows) > 1:  # Mark only if duplicate
            for row_idx in rows:
                ws.cell(row=row_idx, column=review_id_idx).fill = red_fill


    # ✅ Highlight blank cells in Yellow
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_idx in range(1, len(fields) + 1):
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value in [None, "", False, "False", 0, "0", "Unknown", "unknown"]:
                cell.fill = yellow_fill

    # ✅ Highlight invalid URLs
    url_columns = ["link", "reviewImagesUrl"]
    for url_col in url_columns:
        col_idx = fields.index(url_col) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and not str(cell.value).startswith("http"):
                cell.fill = yellow_fill

    # ✅ Highlight status codes that are not 200
    status_col_idx = fields.index("statuscode") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=status_col_idx)
        if cell.value and str(cell.value) != "200":
            cell.fill = red_fill  

    # ✅ Highlight headers in Red if all non-null values are the same
    for col_idx, header in enumerate(fields, start=1):
        column_values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(2, ws.max_row + 1)]
        non_empty_values = [val for val in column_values if val not in [None, "", " "]]
        
        if non_empty_values and len(set(non_empty_values)) == 1:
            ws.cell(row=1, column=col_idx).fill = red_fill

    wb.save(output_path)
    print(f"Reviews saved to {output_path}")

def main():
    # ✅ Edit these values easily
    root_domain = "hp.com/us"
    skus = [
  "HP-LAPTOP-17-CP3047NR",
  "HP-LAPTOP-17T-CN400-173-9Z462AV-1",
  "HP-LAPTOP-17-CN4047NR",
  "HP-LAPTOP-17T-CN200-4V794AV-1",
  "HP-LAPTOP-17T-CN300-173-7P3Q0AV-1",
  "HP-ENVY-INSPIRE-7955E-ALL-IN-ONE-PRINTER",
  "HP-LAPTOP-17T-CN400-173-978Y0AV-1",
  "HP-OFFICEJET-8015E-ALL-IN-ONE-PRINTER",
  "HP-LAPTOP-17Z-CP300-799V0AV-1",
  "HP-PAVILION-LAPTOP-15-EH3047NR",
  "HP-ENVY-INSPIRE-7255E-ALL-IN-ONE-PRINTER",
  "HP-SERIES-5-27-INCH-FHD-MONITOR-WITH-SPEAKERS-HP-WIRELESS-MOUSE-AND-KEYBOARD-HP-FHD-WEBCAM-BUNDLE-527SA-KB-MS-CAM-KIT",
  "HP-SERIES-5-27-INCH-FHD-MONITOR-WITH-SPEAKERS-HP-WIRELESS-MOUSE-AND-KEYBOARD-HP-FHD-WEBCAM-BUNDLE-527SA-KB-MS-CAM-KIT",
  "HP-M27FQ-QHD-MONITOR",
  "HP-SERIES-5-27-INCH-FHD-WHITE-MONITOR-HP-WIRELESS-MOUSE-AND-KEYBOARD-300-BUNDLE-527SW-KB-MS-KIT",
  "HP-V27I-G5-FHD-MONITOR",
  "HP-SERIES-5-238-INCH-FHD-MONITOR-HP-WIRELESS-MOUSE-AND-KEYBOARD-300-BUNDLE-524SH-KB-MS-KIT"
] # Add multiple SKUs here
    max_pages = 4  # Number of pages per SKU
    retry_attempts = 3  # Max retry attempts for API requests
    timeout = 5000 # Timeout for API requests in seconds
    workers = 100  # Number of workers for ThreadPoolExecutor

    output_path = fr"Reviews_for_{root_domain.replace('.', '_').replace('/','-')}.xlsx"
    
    all_reviews = []
    
    with ThreadPoolExecutor(workers) as executor:
        future_to_sku = {executor.submit(fetch_reviews, root_domain, sku, max_pages, retry_attempts, timeout): sku for sku in skus}
        
        for future in as_completed(future_to_sku):
            sku = future_to_sku[future]
            try:
                reviews = future.result()
                all_reviews.extend(reviews)
            except Exception as e:
                print(f"Error processing SKU {sku}: {e}")

    if all_reviews:
        save_to_excel(all_reviews, output_path)
    else:
        print("No reviews retrieved. Excel file will not be created.")

if __name__ == "__main__":
    main()
