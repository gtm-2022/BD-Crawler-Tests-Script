import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed
import datetime

# Function to fetch data from a URL
def get_data_from_url(timeout,isapi,rootdomain,url):
    try:
        if isinstance(url, float):  
            url = str(int(url))  # Remove decimal part and convert to string  
        else:  
            url = str(url).lower()  # Con
        response=""
        if isapi:
            response = requests.get(f"http://localhost:5000/api/apiextraction?rootdomain={rootdomain}&skus={url}", timeout=timeout)
            count = 1

            if response.status_code != 200:
                response = requests.get(f"http://localhost:5000/api/apiextraction?rootdomain={rootdomain}&skus={url}", timeout=timeout*2)
                count += 1

            if response.status_code != 200:
                response = requests.get(f"http://localhost:5000/api/apiextraction?rootdomain={rootdomain}&skus={url}", timeout=timeout*3)
                count += 1
        else:
            response = requests.get(f"http://localhost:5000/api/extraction/sku?rootdomain={rootdomain}&sku={url}", timeout=timeout)
            count = 1

            if response.status_code != 200:
                response = requests.get(f"http://localhost:5000/api/extraction/sku?rootdomain={rootdomain}&sku={url}", timeout=timeout*2)
                count += 1

            if response.status_code != 200:
                response = requests.get(f"http://localhost:5000/api/extraction/sku?rootdomain={rootdomain}&sku={url}", timeout=timeout*3)
                count += 1

        if response.status_code != 200:
            print(f"{url} not worked after {count} attempts")
            return {"status_code": response.status_code, "error": f"Failed after {count} attempts", "sku": url,}
        if response.status_code == 200:
            data = response.json()
            print(f"{url} worked after {count} attempts")

            sellerSku=""
            if isapi:
                sellerSku=data.get("sellerSkus", {})[0]
            else:
                sellerSku = data.get("sellerSku", {})
            skuEntry = sellerSku.get("skuEntry", {})
            skuImages = skuEntry.get("skuImages", {})
            shipping_options = skuEntry.get("shippingOptions", [])

            image_urls = [
                skuImages.get(f'productImageUrl{i}') for i in range(1, 11)
            ]

            # Check if all image URLs start with "http" or "https"
            images_start_with_http = all(url is None or str(url).startswith("http") for url in image_urls)

            return {
                "status_code": response.status_code,
                "price_skuSeller": sellerSku.get("price"),  # Price from skuSeller
                "price_skuEntry": skuEntry.get("price"),  # Price from skuEntry
                "price": sellerSku.get("price") or skuEntry.get("price"),
                "condition": sellerSku.get("condition"),
                "source": sellerSku.get("source"),
                "sku": skuEntry.get("sku"),
                "url": skuEntry.get("url"),
                "name": skuEntry.get("name"),
                "brand": skuEntry.get("brand"),
                "description": skuEntry.get("description"),
                "features": skuEntry.get("features"),
                "upc": skuEntry.get("upc"),
                "ean": skuEntry.get("ean"),
                "mpn": skuEntry.get("mpn"),
                "item_number": skuEntry.get("itemNumber"),
                "store_sku": skuEntry.get("storeSku"),
                "store_name": skuEntry.get("storeName"),
                "availability": skuEntry.get("availability"),
                "category": skuEntry.get("category"),
                "attributes": skuEntry.get("attributes"),
                "star_rating_distribution": skuEntry.get("starRatingDistribution"),
                "average_customer_review": skuEntry.get("averageCustomerReview"),
                "number_of_customer_reviews": skuEntry.get("numberOfCustomerReviews"),
                "variants": skuEntry.get("variants"),
                "parent_sku": skuEntry.get("parentSku"),
                "seller_name": skuEntry.get("sellerName"),
                "seller_id": sellerSku.get("skuEntry", {}).get("buyBoxWinnerHistory"),
                "quantity_sold": skuEntry.get("quantitySold"),
                "quantity_sold_7d": skuEntry.get("quantitySold7D"),
                "variant_attributes": skuEntry.get("variantAttributes"),
                "number_of_favorites": skuEntry.get("numberOfFavorites"),
                "deal_type": skuEntry.get("dealType"),
                "deal_text": skuEntry.get("dealText"),
                "promo_text": skuEntry.get("promoText"),
                "list_price": skuEntry.get("listPrice"),
                "numberOfPayments" : skuEntry.get('numberOfPayments'),
                "pricePerPayments": skuEntry.get('pricePerPayments'),
                'totalPaymentsPrice': skuEntry.get('totalPaymentsPrice'),
                "productImageUrl1": skuImages.get('productImageUrl1'),
                "productImageUrl2": skuImages.get('productImageUrl2'),
                "productImageUrl3": skuImages.get('productImageUrl3'),
                "productImageUrl4": skuImages.get('productImageUrl4'),
                "productImageUrl5": skuImages.get('productImageUrl5'),
                "productImageUrl6": skuImages.get('productImageUrl6'),
                "productImageUrl7": skuImages.get('productImageUrl7'),
                "productImageUrl8": skuImages.get('productImageUrl8'),
                "productImageUrl9": skuImages.get('productImageUrl9'),
                "productImageUrl10": skuImages.get('productImageUrl10'),
                "Images_starts_with_http": images_start_with_http,
                "product_Image_match": skuImages.get('productImageUrl1') == skuImages.get('productImageUrl2') == skuImages.get('productImageUrl3') ==skuImages.get('productImageUrl4') == skuImages.get('productImageUrl5') == skuImages.get('productImageUrl6') == skuImages.get('productImageUrl7') == skuImages.get('productImageUrl8') == skuImages.get('productImageUrl9') == skuImages.get('productImageUrl10'),
                "used_price": skuEntry.get("usedPrice"),
                "model": skuEntry.get("model"),
                "image_count": skuEntry.get("imageCount"),
                "video_count": skuEntry.get("videoCount"),
                "document_count": skuEntry.get("documentCount"),
                "isSponsored": skuEntry.get("isSponsored"),
                "coupon_absolute_discount": skuEntry.get("couponAbsoluteDiscount"),
                "coupon_percent_discount": skuEntry.get("couponPercentDiscount"),
                "panorama_count": skuEntry.get("panoramaCount"),
                "is_aplus": skuEntry.get("isAPlus"),
                "aplus_premium": skuEntry.get("aplusPremium"),
                "aplus_comparison": skuEntry.get("aplusComparison"),
                "aplus_faq": skuEntry.get("aplusFaq"),
                "aplus_video": skuEntry.get("aplusVideo"),
                "flash_sale_end_time": skuEntry.get("flashSaleEndTime"),
                "is_official_seller": skuEntry.get("isOfficialSeller"),
                "price_by_unit": skuEntry.get("priceByUnit"),
                "price_per_unit": skuEntry.get("pricePerUnit"),
                "currency": skuEntry.get("currency"),
                "uvp": skuEntry.get("uvp"),
                "shipping_options": shipping_options,
                "process_name": "cds",
                "timestamp": datetime.datetime.now().isoformat(),
                "rootdomain": skuEntry.get("rootDomain"),
                "preorder": "",
                "category_l1": skuEntry.get("categoryLvl1"),
                "category_l2": skuEntry.get("categoryLvl2"),
                "category_l3": skuEntry.get("categoryLvl3"),
                "category_l4": skuEntry.get("categoryLvl4"),
                "category_l5": skuEntry.get("categoryLvl5"),
                "category_l6": skuEntry.get("categoryLvl6"),
                "category_l7": skuEntry.get("categoryLvl7"),
                "category_l8": skuEntry.get("categoryLvl8"),
                "category_l9": skuEntry.get("categoryLvl9"),
                "category_l10": skuEntry.get("categoryLvl10"),
                "normalized_attributes": "",
                "title_attributes": "",
                "tagged_name": "",
                "number_of_customer_ratings": skuEntry.get("numberOfCustomerRatings"),
                "redirected_sku": skuEntry.get("redirectedSku"),
            }
        else:
            return None
    except requests.RequestException as e:
        print(f"Error fetching data from {url}: {e}")
        return None

# Function to read URLs from the Excel file
def read_urls_from_excel(file_path, column_name='sku'):
    df = pd.read_excel(file_path)
    return df[column_name].tolist()

# Function to write data to a new Excel file and highlight empty columns
def write_data_to_excel(output_file, data_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Define headers
    headers = ["status_code","error",
        "price", "condition","source","sku", "url", "name", "brand", "description", "features",
        "upc", "ean", "mpn", "item_number", "store_sku","store_name", "availability", "category",
        "attributes","star_rating_distribution", "average_customer_review", "number_of_customer_reviews", "variants",
        "parent_sku", "seller_name", "seller_id", "quantity_sold",
        "quantity_sold_7d", "variant_attributes", "number_of_favorites", "deal_type",
        "deal_text", "promo_text", "list_price","numberOfPayments", "pricePerPayments","totalPaymentsPrice",
        "productImageUrl1","productImageUrl2","productImageUrl3","productImageUrl4","productImageUrl5","productImageUrl6",
        "productImageUrl7","productImageUrl8","productImageUrl9","productImageUrl10","Images_starts_with_http","product_Image_match","used_price","model",
        "image_count", "video_count","document_count","isSponsored", "coupon_absolute_discount", "coupon_percent_discount", "panorama_count",
        "is_aplus", "aplus_premium", "aplus_comparison", "aplus_faq", "aplus_video",
        "flash_sale_end_time", "is_official_seller", "price_by_unit", "price_per_unit",
        "currency", "uvp", "shipping_options", "process_name", "timestamp", "rootdomain",
        "preorder", "category_l1", "category_l2", "category_l3", "category_l4", "category_l5",
        "category_l6", "category_l7", "category_l8", "category_l9", "category_l10",
        "normalized_attributes", "title_attributes", "tagged_name", "number_of_customer_ratings",
        "redirected_sku"
    ]
    ws.append(headers)

    for data in data_list:
        row = []
        for field in headers:
            value = data.get(field, "")

            # Convert lists or dictionaries to strings
            if isinstance(value, (list, dict)):
                value = str(value)

            row.append(value)
        ws.append(row)


    # Load workbook for formatting
    wb.save(output_file)
    wb = load_workbook(output_file)
    ws = wb.active

    # Identify and highlight empty columns
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        column_values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(2, ws.max_row + 1)]
    
        # Remove None values to avoid false matches
        non_empty_values = [val for val in column_values if val not in [None, "", False,"False",0,"0","Unknown","unknown"]]

        if len(set(non_empty_values)) == 1 and len(non_empty_values) > 0:  # Check if all values are the same
            ws.cell(row=1, column=col_idx).fill = red_fill  # Highlight header in red

    for col_idx, header in enumerate(headers, start=1):
        is_column_blank = True
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value not in [None, ""]:
                is_column_blank = False
            elif header in headers:
                cell.fill = highlight_fill
        if is_column_blank:
            ws.cell(row=1, column=col_idx).fill = highlight_fill

        for row_idx in range(2, ws.max_row + 1):
            product_Image_match_cell = ws.cell(row=row_idx, column=headers.index("product_Image_match") + 1)
            if product_Image_match_cell.value == True:
                product_Image_match_cell.fill = highlight_fill

        for row_idx in range(2, ws.max_row + 1):
            http_Image_match_cell = ws.cell(row=row_idx, column=headers.index("Images_starts_with_http") + 1)
            if http_Image_match_cell.value == False:
                http_Image_match_cell.fill = highlight_fill

        for row_idx in range(2, ws.max_row + 1):
            availability_cell = ws.cell(row=row_idx, column=headers.index("availability") + 1)
            if availability_cell.value not in [1,"1"]:
                availability_cell.fill = highlight_fill

        for row_idx in range(2, ws.max_row + 1):
            condition_col_idx = headers.index("condition") + 1  # Find the column index for "condition"
            condition_cell = ws.cell(row=row_idx, column=condition_col_idx)
            
            if condition_cell.value not in [1, "1"]:  # Check if the value is not 1
                condition_cell.fill = highlight_fill  # Highlight the cell
        
        for row_idx in range(2, ws.max_row + 1):
            attributes_cell = ws.cell(row=row_idx, column=headers.index("attributes") + 1)
            if isinstance(attributes_cell.value, str):  # Ensuring it's a string (from JSON conversion)
                try:
                    attributes_dict = eval(attributes_cell.value)  # Convert string back to dictionary
                    if isinstance(attributes_dict, dict):
                        for key, value in attributes_dict.items():
                            if value in [None, ""]:  # If any key-value pair is null or empty
                                attributes_cell.fill = highlight_fill
                                break  # Highlight once and stop checking further
                except:
                    pass  # Ignore any conversion errors

        for row_idx in range(2, ws.max_row + 1):
            for img_col in ["productImageUrl1", "productImageUrl2", "productImageUrl3", "productImageUrl4", "productImageUrl5",
                            "productImageUrl6", "productImageUrl7", "productImageUrl8", "productImageUrl9", "productImageUrl10"]:
                col_idx = headers.index(img_col) + 1  # Find the column index
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and not str(cell.value).startswith("http" or "https"):  # Check if URL is invalid
                    cell.fill = highlight_fill  # Highlight the cell


    wb.save(output_file)
    # wb.save(output_file)
# Main function
def main(timeout,isapi,rootdomain,workers,input_file, output_file):
    urls = read_urls_from_excel(input_file)
    extracted_data = []

    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_url = {executor.submit(get_data_from_url,timeout,isapi,rootdomain, url): url for url in urls}
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                result = future.result()
                if result:
                    extracted_data.append(result)
                else:
                    print(f"No data for {url}")
            except Exception as e:
                print(f"Error processing {url}: {e}")

    write_data_to_excel(output_file, extracted_data)

# Run the program
if __name__ == "__main__":
    rootdomain = "hp.com/us"
    timeout=50000
    workers=100
    isapi=True
    #input_excel = fr"Search_{rootdomain.replace('.','_').replace('/','_')}.xlsx"
    input_excel=fr"Search_hp_com_us.xlsx"
    output_excel = fr"pdp_{rootdomain.replace(".","-").replace("/","-")}.xlsx"
    main(timeout,isapi,rootdomain,workers,input_excel, output_excel)